"""Exportación de los metadatos de las capas a hojas de cálculo Excel.

Produce dos archivos con el MISMO contenido en el directorio base:
  - metadatos_capas_campos_en_filas.xlsx     (campos en filas, capas en columnas)
  - metadatos_capas_campos_en_columnas.xlsx  (campos en columnas, capas en filas)

Descubre los XML en ``<base>/<nodo>/3-actual/*.xml`` (nueva convención), con
retroceso a ``<base>/<nodo>/actual/*.xml``; excluye respaldos ``bak*.xml``.
La etiqueta de cada capa es el nombre del archivo.
"""

from __future__ import annotations

from pathlib import Path


def L(n):
    return f"*[local-name()='{n}']"


def first(root, xp):
    """Une (sin duplicados) los textos/atributos que devuelve un xpath."""
    vals = []
    for s in root.xpath(xp):
        if not isinstance(s, str):
            s = s.text or ""
        s = " ".join(s.split())
        if s and s not in vals:
            vals.append(s)
    return "; ".join(vals)


def path_text(p):
    """xpath a //a/b/c/text() a partir de 'a/b/c'."""
    return "//" + "/".join(L(e) for e in p.split("/")) + "/text()"


def path_attr(p, a="codeListValue"):
    return "//" + "/".join(L(e) for e in p.split("/")) + f"/@{a}"


def date_by_type(scope, t):
    return (f"//{L(scope)}//{L('CI_Date')}"
            f"[{L('dateType')}/*[@codeListValue='{t}']]/{L('date')}/*/text()")


# Bases de los dos contactos
PIC = f"//{L('identificationInfo')}//{L('pointOfContact')}[1]"
AUT = f"/*[local-name()='MD_Metadata']/{L('contact')}[1]"

# Sub-campos de contacto: (etiqueta, xpath relativo a la base)
CONTACT_SUBS = [
    ("Rol", f"/{L('CI_Responsibility')}/{L('role')}/{L('CI_RoleCode')}/@codeListValue"),
    ("Organización", f"//{L('CI_Organisation')}/{L('name')}/{L('CharacterString')}/text()"),
    ("Persona", f"//{L('CI_Individual')}/{L('name')}/{L('CharacterString')}/text()"),
    ("Cargo", f"//{L('positionName')}/{L('CharacterString')}/text()"),
    ("Teléfono", f"//{L('number')}/{L('CharacterString')}/text()"),
    ("Correo", f"//{L('electronicMailAddress')}/{L('CharacterString')}/text()"),
    ("Dirección", f"//{L('deliveryPoint')}/{L('CharacterString')}/text()"),
    ("Ciudad", f"//{L('city')}/{L('CharacterString')}/text()"),
    ("Provincia/Área", f"//{L('administrativeArea')}/{L('CharacterString')}/text()"),
    ("Código postal", f"//{L('postalCode')}/{L('CharacterString')}/text()"),
    ("País", f"//{L('country')}/{L('CharacterString')}/text()"),
    ("Horario de atención", f"//{L('hoursOfService')}/{L('CharacterString')}/text()"),
]

# Campos escalares: (etiqueta, xpath). El loop usa first(root, xpath).
SCALAR = [
    ("Identificador del metadato", path_text("metadataIdentifier/MD_Identifier/code/CharacterString")),
    ("Espacio de código", path_text("metadataIdentifier/MD_Identifier/codeSpace/CharacterString")),
    ("Idioma", path_attr("LanguageCode")),
    ("País (ISO 3166)", path_attr("CountryCode")),
    ("Codificación", path_attr("MD_CharacterSetCode")),
    ("Alcance del metadato", path_attr("metadataScope/MD_MetadataScope/resourceScope/MD_ScopeCode")),
    ("Fecha de creación (metadato)", date_by_type("MD_Metadata", "creation")),
    ("Estándar del metadato", path_text("metadataStandard/CI_Citation/title/CharacterString")),
    ("Versión del estándar", path_text("metadataStandard/CI_Citation/edition/CharacterString")),
    ("Frecuencia mant. (metadato)", f"//{L('metadataMaintenance')}//{L('MD_MaintenanceFrequencyCode')}/@codeListValue"),
    # Representación espacial
    ("Nivel de topología", path_attr("MD_TopologyLevelCode")),
    ("Tipo de geometría", path_attr("MD_GeometricObjectTypeCode")),
    ("Nº de geometrías", path_text("geometricObjectCount/Integer")),
    # Sistema de referencia
    ("Sistema de referencia (CRS)", path_text("referenceSystemIdentifier/MD_Identifier/code/CharacterString")),
    ("Tipo de sistema de referencia", path_attr("MD_ReferenceSystemTypeCode")),
    # Identificación del recurso
    ("Título", path_text("MD_DataIdentification/citation/CI_Citation/title/CharacterString")),
    ("Título alternativo", path_text("MD_DataIdentification/citation/CI_Citation/alternateTitle/CharacterString")),
    ("Fecha de publicación (recurso)", date_by_type("MD_DataIdentification", "publication")),
    ("Fecha de edición", path_text("MD_DataIdentification/citation/CI_Citation/editionDate/DateTime")),
    ("Edición", path_text("MD_DataIdentification/citation/CI_Citation/edition/CharacterString")),
    ("Forma de presentación", path_attr("CI_PresentationFormCode")),
    ("Resumen", path_text("abstract/CharacterString")),
    ("Propósito", path_text("purpose/CharacterString")),
    ("Créditos", path_text("credit/CharacterString")),
    ("Estado", path_attr("status/MD_ProgressCode")),
    ("Tipo de representación espacial", path_attr("MD_SpatialRepresentationTypeCode")),
    ("Escala (denominador)", path_text("denominator/Integer")),
    ("Categoría temática", path_text("topicCategory/MD_TopicCategoryCode")),
    ("Frecuencia mant. (recurso)", f"//{L('resourceMaintenance')}//{L('MD_MaintenanceFrequencyCode')}/@codeListValue"),
    # Extensión geográfica
    ("Extensión Oeste", path_text("westBoundLongitude/Decimal")),
    ("Extensión Este", path_text("eastBoundLongitude/Decimal")),
    ("Extensión Sur", path_text("southBoundLatitude/Decimal")),
    ("Extensión Norte", path_text("northBoundLatitude/Decimal")),
    # Palabras clave
    ("Palabras clave", path_text("descriptiveKeywords/MD_Keywords/keyword/CharacterString")),
    ("Tipos de palabra clave", path_attr("MD_KeywordTypeCode")),
    # Restricciones
    ("Restricciones de acceso", path_attr("accessConstraints/MD_RestrictionCode")),
    ("Restricciones de uso", path_attr("useConstraints/MD_RestrictionCode")),
    ("Otras restricciones", path_text("otherConstraints/CharacterString")),
    ("Limitación de uso", path_text("useLimitation/CharacterString")),
    # Información suplementaria
    ("Información suplementaria", path_text("supplementalInformation/CharacterString")),
    # Vista previa
    ("Imagen de vista previa (fileName)", path_text("graphicOverview/MD_BrowseGraphic/fileName/CharacterString")),
    ("Descripción de la imagen", path_text("graphicOverview/MD_BrowseGraphic/fileDescription/CharacterString")),
    # Distribución
    ("Formato de distribución", path_text("distributionFormat/MD_Format/formatSpecificationCitation/CI_Citation/title/CharacterString")),
    ("URL del servicio (distribución)", path_text("transferOptions/MD_DigitalTransferOptions/onLine/CI_OnlineResource/linkage/CharacterString")),
    ("Protocolo (distribución)", path_text("transferOptions/MD_DigitalTransferOptions/onLine/CI_OnlineResource/protocol/CharacterString")),
    ("Nombre de la capa (distribución)", path_text("transferOptions/MD_DigitalTransferOptions/onLine/CI_OnlineResource/name/CharacterString")),
    ("Descripción del recurso en línea", path_text("transferOptions/MD_DigitalTransferOptions/onLine/CI_OnlineResource/description/CharacterString")),
    ("Función del recurso en línea", path_attr("CI_OnLineFunctionCode")),
    # Calidad y linaje
    ("Nivel de calidad (scope)", path_attr("dataQualityInfo/DQ_DataQuality/scope/MD_Scope/level/MD_ScopeCode")),
    ("Linaje", path_text("LI_Lineage/statement/CharacterString")),
]

# Campos cuyo contenido es largo -> columnas anchas en la vista "campos en columnas".
LONG = {"Título", "Título alternativo", "Resumen", "Propósito", "Créditos",
        "Información suplementaria", "Linaje", "Palabras clave", "Otras restricciones",
        "Limitación de uso", "URL del servicio (distribución)",
        "Imagen de vista previa (fileName)", "Descripción de la imagen",
        "Descripción del recurso en línea", "Dirección"}


def construir_campos():
    """Lista final de (etiqueta, xpath). Empieza con Archivo, luego escalares,
    autor del metadato, punto de contacto."""
    campos = [("Archivo", None)]  # None -> nombre de archivo
    campos += SCALAR
    for sub, rel in CONTACT_SUBS:
        campos.append((f"Autor del metadato — {sub}", AUT + rel))
    for sub, rel in CONTACT_SUBS:
        campos.append((f"Punto de contacto — {sub}", PIC + rel))
    return campos


def aplicar_estilo(ws, nrows, ncols):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    hfill = PatternFill("solid", fgColor="1F4E78")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    top = Alignment(vertical="top", wrap_text=True)
    for j in range(1, ncols + 1):
        c = ws.cell(1, j)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    for i in range(1, nrows + 1):
        for j in range(1, ncols + 1):
            cell = ws.cell(i, j)
            cell.border = border
            if i > 1:
                cell.alignment = top
            if j == 1 and i > 1:
                cell.font = Font(bold=True, size=10)
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False


def descubrir_xml(base: Path, nodo: str | None = None) -> list[Path]:
    """XML de metadatos por nodo: <base>/<nodo>/3-actual/ con retroceso a actual/."""
    capas: list[Path] = []
    nodos = [base / nodo] if nodo else sorted(p for p in base.iterdir() if p.is_dir())
    for dir_nodo in nodos:
        for sub in ("3-actual", "actual"):
            xmls = sorted(
                p for p in (dir_nodo / sub).glob("*.xml")
                if not p.name.lower().startswith("bak")
            )
            if xmls:
                capas.extend(xmls)
                break
    return capas


def generar_excel(base: Path, nodo: str | None = None) -> list[Path]:
    """Genera los dos Excel en `base`. Retorna las rutas escritas."""
    import openpyxl
    from lxml import etree
    from openpyxl.utils import get_column_letter

    base = Path(base)
    if not base.is_dir():
        raise FileNotFoundError(f"No existe el directorio: {base}")

    capas = descubrir_xml(base, nodo)
    if not capas:
        raise FileNotFoundError(f"No se encontraron XML en {base}/*/3-actual/ ni */actual/")

    campos = construir_campos()
    roots = [(p.stem, etree.parse(str(p)).getroot(), p) for p in capas]

    data = {}
    for label, xp in campos:
        data[label] = {}
        for clabel, root, path in roots:
            if xp is None:
                data[label][clabel] = path.name
            else:
                try:
                    data[label][clabel] = first(root, xp)
                except Exception as e:
                    data[label][clabel] = f"(err: {e})"

    out_filas = base / "metadatos_capas_campos_en_filas.xlsx"
    out_cols = base / "metadatos_capas_campos_en_columnas.xlsx"

    # Archivo 1: campos en FILAS, capas en columnas
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metadatos"
    ws.cell(1, 1, "Campo")
    for j, (clabel, _, _) in enumerate(roots, start=2):
        ws.cell(1, j, clabel)
    for i, (label, _) in enumerate(campos, start=2):
        ws.cell(i, 1, label)
        for j, (clabel, _, _) in enumerate(roots, start=2):
            ws.cell(i, j, data[label][clabel])
    aplicar_estilo(ws, len(campos) + 1, len(roots) + 1)
    ws.column_dimensions["A"].width = 34
    for j in range(2, len(roots) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 44
    wb.save(out_filas)

    # Archivo 2: campos en COLUMNAS, capas en filas
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metadatos"
    ws.cell(1, 1, "Capa")
    for j, (label, _) in enumerate(campos, start=2):
        ws.cell(1, j, label)
    for i, (clabel, _, _) in enumerate(roots, start=2):
        ws.cell(i, 1, clabel)
        for j, (label, _) in enumerate(campos, start=2):
            ws.cell(i, j, data[label][clabel])
    aplicar_estilo(ws, len(roots) + 1, len(campos) + 1)
    ws.column_dimensions["A"].width = 40
    for j, (label, _) in enumerate(campos, start=2):
        ws.column_dimensions[get_column_letter(j)].width = 55 if label in LONG else 22
    wb.save(out_cols)

    print(f"Capas: {len(roots)} | Campos: {len(campos)}")
    for clabel, _, _ in roots:
        print(f"  - {clabel}")
    print(f"OK -> {out_filas}")
    print(f"OK -> {out_cols}")
    return [out_filas, out_cols]
